#!/usr/bin/env python3

"""
Written by Nick Turner (@nickjvturner@mastodon.social)
This script will find 'the' word doc in the same directory as the script
Convert the contents of 'the' tables into an XLSX file

Initially created in September 2023
Latest meaningful update: 2023-09-05
"""

from pathlib import Path

from docx import Document

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

nl = '\n'


def main():
	docx_files = []

	# Get local file with extension .docx
	for file in sorted(Path.cwd().iterdir()):
		# ignore docx files in directory with $ and OUTPUT in filename
		if file.suffix == '.docx' and all(char not in file.stem for char in ('$', 'OUTPUT')):
			docx_files.append(file)

	for docx in docx_files:
		print(docx.name)
	proceed = input(f'{nl}Proceed with listed files?: (YES/no)')
	if proceed.lower() == 'no':
		exit()

	# print(f'{nl}filename: {file.name}')

	for docx in docx_files:
		print(f'{nl}{docx.name}')

		# Load the Word document
		doc = Document(docx)

		# Create a new Excel workbook
		wb = Workbook()
		ws = wb.active

		# Set the sheet name (you can modify this line to set a custom name)
		ws.title = 'AP Tracker'

		# Initialize row and column counters
		row_num = 1
		col_num = 1

		# Iterate through tables in the Word document
		for table in doc.tables:
			for row in table.rows:
				for cell in row.cells:
					# Add cell content to the Excel worksheet
					ws.cell(row=row_num, column=col_num, value=cell.text)
					col_num += 1
				col_num = 1  # Reset column counter for each row
				row_num += 1

		# Set font size and make row 1 bold
		bold_font = Font(bold=True, size=16)
		for row in ws.iter_rows(min_row=1, max_row=1):
			for cell in row:
				cell.font = bold_font

		# Auto-fit all columns
		for column in ws.columns:
			max_length = 0
			column_letter = get_column_letter(column[0].column)
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(cell.value)
				except Exception:
					pass
			adjusted_width = (max_length + 5)
			ws.column_dimensions[column_letter].width = adjusted_width

		# Save the Excel workbook
		wb.save(docx.stem + 'converted_table.xlsx')


if __name__ == "__main__":
	main()
